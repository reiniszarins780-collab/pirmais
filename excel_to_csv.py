"""Command-line tool for converting Excel workbooks to CSV files.

This script reads an Excel workbook (``.xlsx``) and exports the
selected worksheet to a CSV file.  The first worksheet is used by
default, but any sheet can be chosen with the ``--sheet`` option.

Example usage::

    python excel_to_csv.py darba_lapa.xlsx --sheet "Dati" --output dati.csv

If no output file is provided, the CSV is written next to the input
file and named after the chosen worksheet.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled via runtime error
    raise SystemExit(
        "Bibliotēka 'openpyxl' nav pieejama. Uzinstalē to ar 'pip install openpyxl'."
    ) from exc


def iter_values(rows: Iterable[Iterable[Optional[object]]]) -> Iterable[list[str]]:
    """Convert worksheet rows to lists of string values.

    ``openpyxl`` returns ``None`` for empty cells.  CSV files expect text,
    so ``None`` is replaced with an empty string.  Other values are cast to
    strings to avoid issues with non-text types (e.g. numbers, dates).
    """

    for row in rows:
        yield ["" if value is None else str(value) for value in row]


def convert_excel_to_csv(
    workbook_path: Path,
    worksheet_name: str,
    output_path: Path,
    delimiter: str,
) -> None:
    """Convert the specified worksheet to a CSV file."""

    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    worksheet = workbook[worksheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=delimiter)
        for row in iter_values(worksheet.iter_rows(values_only=True)):
            writer.writerow(row)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pārveido Excel (.xlsx) failu uz CSV formātu.")
    parser.add_argument("input", type=Path, help="Excel (.xlsx) fails, ko pārveidot.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="CSV faila nosaukums. Ja nav norādīts, tiek izmantots darblapas nosaukums.",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        type=str,
        help="Darblapa, ko eksportēt. Noklusēti tiek izmantota pirmā darblapa.",
    )
    parser.add_argument(
        "-d",
        "--delimiter",
        default=",",
        help="CSV atdalītājs (noklusējums ir komats).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    input_path: Path = args.input
    if not input_path.is_file():
        raise SystemExit(f"Fails '{input_path}' netika atrasts.")

    sheet_name: Optional[str] = args.sheet

    workbook = load_workbook(filename=input_path, read_only=True)

    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Darblapa '{sheet_name}' netika atrasta failā '{input_path.name}'.")
        worksheet_name = sheet_name
    else:
        worksheet_name = workbook.active.title

    if args.output is not None:
        output_path = args.output
    else:
        output_path = input_path.with_name(f"{worksheet_name}.csv")

    convert_excel_to_csv(input_path, worksheet_name, output_path, args.delimiter)


if __name__ == "__main__":
    main()
